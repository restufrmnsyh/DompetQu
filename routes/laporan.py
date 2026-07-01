from flask import Blueprint, render_template, request, redirect, send_file, session, flash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
import json
import sqlite3
from database.db import (
    ambil_semua_transaksi, 
    restore_transaksi,
    hitung_ringkasan,
    hitung_saldo,
    reset_transaksi
    )

laporan = Blueprint("laporan", __name__)


@laporan.route("/export")
def export_excel():

    if "login" not in session:
        return redirect("/login")

    transaksi = ambil_semua_transaksi(session["user_id"])

    pemasukan, pengeluaran = hitung_ringkasan(session["user_id"])
    saldo = hitung_saldo(session["user_id"])

    wb = Workbook()

    # ===============================
    # SHEET RINGKASAN
    # ===============================

    ws = wb.active
    ws.title = "Ringkasan"

    ws.merge_cells("A1:E1")
    ws["A1"] = "RINGKASAN KEUANGAN"
    ws["A1"].font = Font(size=18, bold=True, color="1E40AF")

    ws.merge_cells("A2:E2")
    ws["A2"] = f"DompetQu • {session['username']} • {datetime.now().strftime('%d %B %Y')}"
    ws["A2"].font = Font(size=10, italic=True)

    # Card Ringkasan
    data_ringkasan = [
        ("Total Pemasukan", pemasukan),
        ("Total Pengeluaran", pengeluaran),
        ("Saldo Bersih", saldo)
    ]

    warna = [
        "C6F6D5",
        "FEE2E2",
        "DBEAFE"
    ]

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    start_col = 1

    for i, (judul, nilai) in enumerate(data_ringkasan):

        col = start_col + i * 2

        ws.cell(row=5, column=col).value = judul
        ws.cell(row=5, column=col).font = Font(bold=True)
        ws.cell(row=5, column=col).fill = PatternFill(
            "solid",
            fgColor=warna[i]
        )

        ws.cell(row=6, column=col).value = nilai
        ws.cell(row=6, column=col).number_format = '"Rp" #,##0'
        ws.cell(row=6, column=col).font = Font(
            size=15,
            bold=True
        )
        ws.cell(row=6, column=col).fill = PatternFill(
            "solid",
            fgColor=warna[i]
        )

        ws.cell(row=5, column=col).border = border
        ws.cell(row=6, column=col).border = border

    # ===============================
    # SHEET PEMASUKAN
    # ===============================

    ws_masuk = wb.create_sheet("💰 Pemasukan")

    ws_masuk.merge_cells("A1:E1")
    ws_masuk["A1"] = "LAPORAN PEMASUKAN — DompetQu"
    ws_masuk["A1"].font = Font(
        size=18,
        bold=True,
        color="0D7A55"
    )

    ws_masuk["A2"] = f"Diekspor pada {datetime.now().strftime('%d %B %Y %H:%M')}"
    ws_masuk["A2"].font = Font(italic=True)

    header = [
        "No",
        "Tanggal",
        "Kategori",
        "Nominal",
        "Catatan"
    ]

    for i, h in enumerate(header, 1):

        c = ws_masuk.cell(row=5, column=i)

        c.value = h
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0D7A55")
        c.alignment = Alignment(horizontal="center")
        c.border = border

    row = 6
    total = 0

    for no, t in enumerate(transaksi, 1):

        if t[2] != "Pemasukan":
            continue

        ws_masuk.cell(row=row, column=1).value = no
        ws_masuk.cell(row=row, column=2).value = t[1]
        ws_masuk.cell(row=row, column=3).value = t[3]

        nominal = ws_masuk.cell(row=row, column=4)
        nominal.value = int(t[4])
        nominal.number_format = '#,##0'

        ws_masuk.cell(row=row, column=5).value = t[5] or ""

        row += 1

    ws_masuk[f"C{row+1}"] = "TOTAL PEMASUKAN"
    ws_masuk[f"C{row+1}"].font = Font(
        bold=True,
        color="0D7A55"
    )

    ws_masuk[f"D{row+1}"] = total
    ws_masuk[f"D{row+1}"].number_format = '"Rp" #,##0'

    # ===============================
    # SHEET PENGELUARAN
    # ===============================

    ws_keluar = wb.create_sheet("💸 Pengeluaran")

    ws_keluar.merge_cells("A1:E1")
    ws_keluar["A1"] = "LAPORAN PENGELUARAN — DompetQu"
    ws_keluar["A1"].font = Font(
        size=18,
        bold=True,
        color="B91C1C"
    )

    ws_keluar["A2"] = f"Diekspor pada {datetime.now().strftime('%d %B %Y %H:%M')}"
    ws_keluar["A2"].font = Font(italic=True)

    for i, h in enumerate(header, 1):

        c = ws_keluar.cell(row=5, column=i)

        c.value = h
        c.font = Font(
            bold=True,
            color="FFFFFF"
        )
        c.fill = PatternFill(
            "solid",
            fgColor="B91C1C"
        )
        c.alignment = Alignment(horizontal="center")
        c.border = border

    row = 6
    total = 0

    for no, t in enumerate(transaksi, 1):

        if t[2] != "Pengeluaran":
            continue

        ws_keluar.cell(row=row, column=1).value = no
        ws_keluar.cell(row=row, column=2).value = t[1]
        ws_keluar.cell(row=row, column=3).value = t[3]

        nominal = ws_keluar.cell(row=row, column=4)
        nominal.value = int(t[4])
        nominal.number_format = '#,##0'

        ws_keluar.cell(row=row, column=5).value = t[5] or ""

        row += 1

    ws_keluar[f"C{row+1}"] = "TOTAL PENGELUARAN"
    ws_keluar[f"C{row+1}"].font = Font(
        bold=True,
        color="B91C1C"
    )

    ws_keluar[f"D{row+1}"] = total
    ws_keluar[f"D{row+1}"].number_format = '"Rp" #,##0'

    # ===============================
    # COLUMN WIDTHS
    # ===============================

    from openpyxl.utils import get_column_letter

    for sheet in wb.worksheets:
        for column_cells in sheet.iter_cols():

            max_length = 0
            column = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                try:
                    if cell.value and not isinstance(cell.value, (int, float)):
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            sheet.column_dimensions[column].width = min(max_length + 4, 40)

    # Override manual untuk kolom angka di sheet Pemasukan & Pengeluaran
    for sheet in [ws_masuk, ws_keluar]:
        sheet.column_dimensions["A"].width = 6   # No
        sheet.column_dimensions["B"].width = 14  # Tanggal
        sheet.column_dimensions["C"].width = 20  # Kategori
        sheet.column_dimensions["D"].width = 20  # Nominal — cukup untuk "Rp 99.999.999"
        sheet.column_dimensions["E"].width = 36  # Catatan

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name="laporan_keuangan.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@laporan.route("/backup")
def backup():
    if "login" not in session:
        return redirect("/login")

    conn = sqlite3.connect("database.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM transaksi WHERE user_id = ?", (session["user_id"],))
    data = [dict(row) for row in cursor.fetchall()]
    conn.close()

    output = BytesIO(json.dumps(data, ensure_ascii=False, indent=4).encode("utf-8"))
    output.seek(0)
    return send_file(
        output,
        download_name="backup_dompetku.json",
        as_attachment=True,
        mimetype="application/json"
    )


@laporan.route("/restore", methods=["GET", "POST"])
def restore():
    if "login" not in session:
        return redirect("/login")

    pesan = ""
    if request.method == "POST":
        file = request.files.get("file")
        if file:
            try:
                data = json.load(file)
                restore_transaksi(session["user_id"], data)
                return redirect("/")
            except Exception as e:
                pesan = f"Gagal restore: {str(e)}"

    return render_template("restore.html", pesan=pesan)

@laporan.route("/reset")
def reset():

    if "login" not in session:
        return redirect("/login")

    reset_transaksi(session["user_id"])

    flash("Semua transaksi berhasil dihapus")

    return redirect("/")